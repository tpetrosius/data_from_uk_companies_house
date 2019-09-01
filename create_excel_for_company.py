import openpyxl
from openpyxl import Workbook

def create_excel(company_info, ownership):
	"""Create excel for company."""
	
	#Create filepath
	folder = r'C:\Users\Vartotojas\Desktop\Projects\UK Companies House API'
	filepath = folder + '\\' + company_info['company_name'] + '.xlsx'

	#Create workbook which is named by company name
	wb = Workbook()
	wb.save(filepath)

	#Open create workbook
	wb = openpyxl.load_workbook(filepath)
	sheet = wb.active

	#Create company information page
	sheet_1 = wb.create_sheet("Company Information", 0)
	row = 1
	for key, value in company_info.items():
		sheet_1['A' + str(row)] = key.title()
		sheet_1['B' + str(row)] = value.title()
		row += 1
		
	#Create ownership page
	sheet_2 = wb.create_sheet("Ownership", 1)

	if len(ownership) != 0:

		#Create columns in ownership page
		column_number = 1
		for key in ownership[1].keys():
			column_name = key.title()
			cell_ref = sheet_2.cell(row=1, column=column_number)
			cell_ref.value = column_name
			column_number += 1
		
		#Write ownership data
		column_number = 1
		for i in range(1, len(ownership)+1):
			for value in ownership[i].values():
				cell_ref = sheet_2.cell(row=i+1, column=column_number)
				cell_ref.value = value.title()
				column_number += 1
			column_number = 1
	else:
		sheet_2['A1'] = 'No information on ownership'

	if 'Sheet' in wb.sheetnames:
		del wb['Sheet']

	wb.save(filepath)
